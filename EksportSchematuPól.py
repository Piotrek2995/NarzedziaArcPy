import arcpy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import os

# --------------------------------
# Parametr wejściowy z ArcGIS (Excel)
# --------------------------------
excel_path = arcpy.GetParameterAsText(0)
aprx = arcpy.mp.ArcGISProject("CURRENT")
m = aprx.activeMap

# --------------------------------
# Przygotuj nowy plik Excel
# --------------------------------
wb = openpyxl.Workbook()
legend_ws = wb.active
legend_ws.title = "Legenda"

legend_ws["A1"] = "Legenda eksportu właściwości pól"
legend_ws["A1"].font = Font(bold=True)

legend_text = [
    ("Kolumny:", ""),
    ("Name", "Nazwa pola w bazie"),
    ("Alias", "Alias pola (najpierw z CIM, jeśli brak to z bazy)"),
    ("Visible", "Widoczność pola w tabeli atrybutów (True/False)"),
    ("ReadOnly", "Czy pole jest tylko do odczytu (True/False)"),
    ("Domain", "Domena przypisana w bazie danych"),
]

row = 3
for name, desc in legend_text:
    legend_ws[f"A{row}"] = name
    legend_ws[f"B{row}"] = desc
    if not name.strip() or name.endswith(":"):
        legend_ws[f"A{row}"].font = Font(bold=True)
    row += 1



default_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

# --------------------------------
# Dodanie informacji o warstwach bez CIM w A11:F22
# --------------------------------
legend_ws.merge_cells("A11:F22")
legend_ws["A11"] = (
    "⚠️ Niektóre warstwy mogą nie przechowywać właściwości CIM.\n\n"
    "W Excelu **nie pojawi się tabela pól dla tych warstw**.\n"
    "Aby włączyć CIM dla warstwy w ArcGIS Pro, należy:\n"
    "1. Otworzyć tabelę atrybutów warstwy.\n"
    "2. Ręcznie zmienić co najmniej jedną właściwość pola "
    "(np. alias, widoczność).\n"
    "3. Zapisz projekt/warstwę. Od tego momentu warstwa będzie przechowywać CIM, "
    "a eksport do Excela pokaże tabelę pól."
)
legend_ws["A11"].alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
legend_ws["A11"].font = Font(color="FF0000", bold=True)

def tf_str(val):
    if isinstance(val, bool):
        return "True" if val else "False"
    return val

system_fields = {"OBJECTID", "Shape", "Shape_Length", "Shape_Area"}

# --------------------------------
# Tworzenie arkuszy warstw i zbieranie używanych domen
# --------------------------------
used_domains = {}  # {domena: typ pola używającego jej}

for lyr in m.listLayers():
    if lyr.isFeatureLayer:
        cim_fields_map = {}
        try:
            cim_lyr = lyr.getDefinition("V2")
            cim_table = cim_lyr.featureTable
            if hasattr(cim_table, "fieldDescriptions"):
                for f in cim_table.fieldDescriptions:
                    cim_fields_map[f.fieldName] = {
                        "alias": getattr(f, "alias", None),
                        "visible": getattr(f, "visible", None),
                        "readOnly": getattr(f, "readOnly", None),
                    }
        except:
            pass

        try:
            dataset_path = lyr.dataSource
            arcpy_fields = arcpy.ListFields(dataset_path)
        except:
            continue

        # Zbieramy domeny (ArcGIS Pro)
        gdb_path = arcpy.Describe(lyr.dataSource).path
        try:
            domains = arcpy.da.ListDomains(gdb_path)
        except:
            domains = []

        for fld in arcpy_fields:
            if fld.domain:
                used_domains[fld.domain] = fld.type

        sheet_name = lyr.name[:31]
        ws = wb.create_sheet(title=sheet_name)

        has_cim = any(
            v is not None
            for props in cim_fields_map.values()
            for v in props.values()
        )

        if not has_cim:
            ws.merge_cells("A1:E5")
            ws["A1"] = (
                "⚠️ Ta warstwa nie przechowuje właściwości CIM.\n\n"
                "W Excelu nie pojawi się tabela pól dla tej warstwy.\n"
                "Aby włączyć CIM, zmień ręcznie właściwość pola w ArcGIS Pro\n"
                "(np. alias, widoczność) i zapisz projekt/warstwę."
            )
            ws["A1"].font = Font(bold=True, color="FF0000")
            ws["A1"].alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
            continue

        headers = ["Name", "Alias", "Visible", "ReadOnly", "Domain", "Type"]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)

        dv_tf = DataValidation(type="list", formula1='"True,False"', allow_blank=False)
        ws.add_data_validation(dv_tf)

        for fld in arcpy_fields:
            cim_props = cim_fields_map.get(fld.name, {})
            alias = cim_props.get("alias") or fld.aliasName
            alias_fill = cim_props.get("alias") is None
            visible = cim_props.get("visible")
            visible = True if visible is None else visible
            visible_fill = cim_props.get("visible") is None
            readonly = cim_props.get("readOnly")
            if readonly is None:
                readonly = True if fld.name in system_fields else False
                readonly_fill = True
            else:
                readonly_fill = False
            row_values = [fld.name, alias, tf_str(visible), tf_str(readonly), fld.domain, fld.type]
            ws.append(row_values)
            row_idx = ws.max_row
            if alias_fill:
                ws.cell(row=row_idx, column=2).fill = default_fill
            if visible_fill:
                ws.cell(row=row_idx, column=3).fill = default_fill
            if readonly_fill:
                ws.cell(row=row_idx, column=4).fill = default_fill
            dv_tf.add(ws.cell(row=row_idx, column=3))
            dv_tf.add(ws.cell(row=row_idx, column=4))

# --------------------------------
# Dostosowanie szerokości kolumn i zamrożenie nagłówków w arkuszach warstw
# --------------------------------
for ws in wb.worksheets:
    if ws.title not in ["Legenda", "Domeny"]:
        ws.freeze_panes = "A2"  # zamrożenie nagłówków
        for col in ws.columns:
            # Znajdź pierwszą zwykłą komórkę w kolumnie, aby dostać literę
            col_letter = None
            for cell in col:
                if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    col_letter = cell.column_letter
                    break
            if col_letter is None:
                continue  # jeśli cała kolumna to scalone komórki, pomiń

            max_length = 0
            for cell in col:
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2  # margines

# --------------------------------
# Tworzenie arkusza Domeny jako drugi arkusz – wszystkie domeny
# --------------------------------
dom_ws = wb.create_sheet(title="Domeny", index=1)
dom_ws["A1"] = "Arkusz Domeny – wszystkie domeny w geobazie"
dom_ws["A2"] = "Kolumna 'Domena' – nazwa domeny"
dom_ws["A3"] = "Kolumna 'Wartości' – wszystkie wartości domeny w jednej komórce (rozdzielone średnikiem)"

for cell in ["A1", "A2", "A3"]:
    dom_ws[cell].font = Font(bold=True)
    dom_ws[cell].alignment = Alignment(wrap_text=True)

# Szerokość kolumny i wysokość wierszy
dom_ws.column_dimensions["A"].width = 33  # ~240 px
for r in range(1, 4):
    dom_ws.row_dimensions[r].height = 100

dom_ws.append(["Domena", "Wartości"])
gdb_paths = list(set([arcpy.Describe(lyr.dataSource).path for lyr in m.listLayers() if lyr.isFeatureLayer]))
all_domains_final = {}
for gdb_path in gdb_paths:
    try:
        domains = arcpy.da.ListDomains(gdb_path)
        for d in domains:
            try:
                if d.domainType == "CodedValue":
                    values = ";".join([str(k) for k in d.codedValues.keys()]) if d.codedValues else ""
                elif d.domainType == "Range":
                    values = f"{d.range[0]} - {d.range[1]}"
                else:
                    values = ""
            except:
                values = ""
            all_domains_final[d.name] = {"values": values}
    except:
        pass

for d, props in all_domains_final.items():
    dom_ws.append([d, props["values"]])

# --------------------------------
# Zapis pliku
# --------------------------------
if os.path.exists(excel_path):
    os.remove(excel_path)

wb.save(excel_path)
arcpy.AddMessage(f"Zapisano do pliku: {excel_path}")